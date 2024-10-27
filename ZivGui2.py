# V01.0 First Ver For Compiling
# Comments Not Deleted
import openpyxl as OP
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime, timedelta
import os
import re
from tkinter import *

#from tkinter.ttk import *
#from tkinter import *
#from tkinter.filedialog import askopenfilename, askdirectory
from tkinter import filedialog 
from tkinter import messagebox
#from tkinter import simpledialog
import random as rn
import shutil

today = datetime.now()
tomorrow = today + timedelta(1)
tomorrowName = (tomorrow.strftime('%A'))
userName = os.getlogin()
output_path = os.getcwd()
output_file = output_path + '\\zivugim' + tomorrowName + '.xlsx' #ONLY XLSX!!!
output_file = output_path + '\\zivugim' + tomorrowName + '.xlsx' #ONLY XLSX!!!
print(output_path)
print(type(output_path))
rows = []
programs = {}
userName = os.getlogin()
CoolExit = [("Goodbye to all that"), ("Hello Goodbye"), ("I am a camera with its shutter open\nquite passive, recording, not thinking"),
("Time to Say Goodbye"), ("Bang, zoom, straight to the moon"),
("Wish me luck as you wave me goodbye"), ("Did I do that?"), ("D'oh!"), ("Here it is, your moment of Zen"), ("I love it when a plan comes together"),
("You are the weakest link, goodbye"), ("No soup for you!"), ("Schwing!"), ("Tell me what you don't like about yourself"), ("The Tribe has spoken"),
("You can kiss that one goodbye"), ("This tape will self-destruct in five seconds"), ("To the Batmobile!"), ("Two thumbs up"), ("I'll be back"), 
("May the force be with you"),("Im king of the world!"),("Yippee Ki Yay, Motherfucker"),("You've got to ask yourself one question:\nDo I feel lucky? Well, do ya, punk?"),("Go ahead, make my day"),("I love the smell of napalm in the morning"),("Hasta la vista, baby"),
("To infinity…and beyond!"), ("Alrighty then!"), ("They call me Mr. Tibbs!"), ("Shall…we…play…a…game?"), ("It’s showtime!"), ("I know kung fu"), ("I have a feeling we're not in Kansas anymore"), ("Why so serious?"),
("Thank you for your cooperation"), ("Beam me up Scotty"), ("exterminate!"), ("Shit just got real"), ("I know you are\nbut what am I"), ("I have come here to chew bubblegum and kick ass\nand I’m all out of bubblegum"), ("Well, la-de-da"),
("I have a feeling we’re not in Kansas anymore"), ("I see dead people"), ("Bazinga!"), ("Cowabunga!"), ("Hoo-ah!"), ("SHOW ME THE MONEY!!!"), ("Yeah baby"), 
("Are we having fun yet?"), ("You got it, dude"), ("Hakuna Matata!"), ("Ssssssssssssmokin!"), ("Gooble gobble, gooble gobble"), ("It’s alive! It’s alive! IT’S ALIVE!!!"), ("My momma always said, Life is like a box of chocolates\nyou never know what you’re gonna get"), 
("Game over"), ("Been there, done that, got the T-shirt"), ("All over again"), ("All done and dusted"), ("If opportunity doesn’t knock, build a door"), ("If you cannot do great things, do small things in a great way"), ("Be who you are and say what you feel, because\nthose who mind don’t matter and those who matter don’t mind"), 
("Bring me a bucket, and I'll show you a bucket!"), ("Job's Done"), ("I Need A Weapon"), ("The Cake Is A Lie!"), ("Thank You! But Our Princess Is In Another Castle!"), ("Finish Him!"), ("Done!", "OK, You go now"), 
("Yippie ka-yay, motherfucker!"), ("Why so serious?"), ("Wohoo!"), ("Let's rock!"), ("Go ahead, make my day"), ("Get back to work, you slacker!"), ("Stand in the ashes of a trillion dead souls, and ask the ghosts\nif honor matters. The silence is your answer"), 
("Okay, I've just gotta concentrate!"), ("Agh, just... I just gotta get it through here..."), ("Okay, you know what? That's close enough.\nJust hold tight"), ("Just say 'Apple'. Classic. Very simple."), ("Simple word. 'Apple'."), ("HA! I knew someone was alive in here"), ("Hello? Anyone in there?"),
("Your destination's probably not going to\ncome meet us here. Is it? So go on"), ("On ya go"), ("Go on"), ("So, once again, just... move along. One small step and everything."), ("Yeah, it's alright. Go ahead."), ("Alright, off you go!"), ("Aggh, see, now I hit that one, I hit that one..."),
("I can't do it if you're watching.\nIf you.... just turn around?"), ("I can't... I can't do it if you're watching. [nervous laugh]"), ("I'm not joking. Could you just turn around for a second?"), ("Ummmm. Yeah, I can't do it if you're watching."), ("Look down. Where am I? Where am I?"), ("On three. Ready? One... Two..."), ("Come on through to the other side"), ("There should be a portal device\non that podium over there"), ("Hey hey! You made it!"),
("Alright, you can turn around now!"), ("Okay, I've decided not to kill you. IF you press the button"), ("Oh! Wow! Good! I didn't think that was going to work"), ("Hello! This is the part where I kill you"), ("Oh! Oh! Did it kill you? Oh,\nthat would be amazing if it killed you"), ("I forbid you to press the button!"), ("Do not press that button!"), ("Ohhhhhh, we just made it! That was close"), ("We're good! Appreciate it!"), ("And off we go"),
("Huh. That was easy"), ("Agh! You're alive! Great!"), ("There. Bing! Perfect. On you go"), ("Ever have that feeling where you’re not sure\nif you’re awake or dreaming? "),
("Follow the white rabbit"), ("This Window will self destruct, ok?"), ("For a moment, nothing happened. Then, \nafter a second or so, nothing continued to happen"),
("DON’T PANIC"), ("Time is an illusion.\nLunchtime doubly so."), ("What is my purpose?"), ("Please call me Eddie if it will help you to relax"), ("If you don’t open that exit hatch this moment\nI shall zap straight off to your major data banks and\nreprogram you with a very large axe, got that?"), ("So long, and thanks for all the fish"),
("I really Cronenberged up the whole place"), ("I'm about to do to you what Limp-\n-Bizkit did to music in the late '90s"), ("We interrupt this program to annoy you\nand make things generally more irritating"), ("The mice will see you now"), ("I could have more fun in cat litter"), ("keep banging the rocks together"), ("I only know as much about myself as my mind can work out under its current conditions.\nAnd its current conditions are not good."),
('Do geese see goD?'),('Step on no petS'),('As I pee, sir, I see PisA'),
('Was it a car or a cat I saW?'),('Taco caT'),('A nut for a jar of tunA'),('Never Odd Or EveN'),
('Another day, another dollar'), ('Mi scuziii'), ('How you doin?'), ('I am cornholio'), ('Keep the change ya filthy animal'), 
('Chop your own wood\nit will warm you twice'), ('Do you wanna get high?'), ('OH MY GOD!\nI think i killed Kenny'), ('No soup for you'), ('I RUN! slower then internet explorer on a dial up connection,\nbut i run!'), ('You got it dude'), ('It takes as much energy to wish as it does to plan'), 
('Jew!'), ('Whateva, I do what I want!'), ('Aw! God-damn it!'), ('Respect ma authoritah!'), ('Wibbly-Wobbly, Timey-Wimey...Stuff\nDONE!'), ('Wer`re not worthy'), ('Resistance is futile'), 
('Member Ghostbusters?'), ('Member Dagobah? Thats where Yoda lives!'), ('Member Yoda?'), ('Member Jurassic Park?'), ('Member Chewbacca?'), ('Welcome to Shitty Wok. Can I take a order prease?'),
('Dont forget to bring a towel!'), ('TIMMAH!'), 
('Member Chewbacca again?'), ('Wubba Lubba Dub Dub!'), ('There is no Spoon' + userName), ('So we meet again' + userName + '...')]
Coool = rn.randrange(0,168)
fill_yellow = PatternFill(start_color='F0E01F', end_color='F0E01F', fill_type='solid')
fill_green = PatternFill(start_color='58F01F', end_color='58F01F', fill_type='solid')
fill_light_green = PatternFill(start_color='04cc82', end_color='04cc82', fill_type='solid')
fill_blue = PatternFill(start_color='1F68F0', end_color='1F68F0', fill_type='solid')
fill_purple = PatternFill(start_color='981FF0', end_color='981FF0', fill_type='solid')
fill_pink = PatternFill(start_color='F01FA4', end_color='F01FA4', fill_type='solid')
fill_orange = PatternFill(start_color='F0881F', end_color='F0881F', fill_type='solid')
fill_red = PatternFill(start_color='FFFFC7CE', end_color='FFFFC7CE', fill_type='solid')
red_font = Font(color='FF9C0006')


class ZivWindow:
	def __init__(self):
		self.root = Tk()
		self.root.title('זיווגים')
		self.root.geometry('400x230')
		self.root.resizable(0, 0)
		self.label1 = Label(self.root, text="בדרכיו נדבק ואף חשובה מצווה מקיים זיווג העושה אדם")
		self.label2 = Label(self.root, text="בעולם הראשון הזיווג את שעשה הקבה של")
		self.label1.pack()
		self.label2.pack()
		self.runbtn = Button(self.root, text="Run", width=50, command=self.run).place(x=30, y=200)
		self.load_files_button = Button(self.root, text='Load Magic File', command=self.load_file)
		self.load_files_button.pack(anchor='nw')
		self.root.mainloop()


	def load_file(self):
		try:
			self.MagicFile = filedialog.askopenfilename(initialdir='', filetypes=[('Only "*.XLSX"', "*.xlsx")])
			self.wb = OP.load_workbook(self.MagicFile)
			self.sheet = self.wb['גיליון1']
			print('bla', self.sheet)
			print('file loaded', self.MagicFile)
			return self.wb, self.sheet, self.MagicFile
		except PermissionError:
			messagebox.showerror("Whoops!", "Close opened xml file\nAnd load again!")
			

			

	def run(self):
		print('main running', self.wb)
		self.max_row_num = self.sheet.max_row
		self.max_row_let = get_column_letter(self.sheet.max_column)
		self.last_row = self.max_row_let + str(self.max_row_num)
		self.last_row_eng_program_name = 'D' + str(self.max_row_num)
		self.last_row_lacart = 'H' + str(self.max_row_num)


		def build_dict(): # iter throgh specific columns in a row 
			print('build_dict')
			print(self.sheet)
			for _row in range(11, self.max_row_num):
				_progName = self.sheet.cell(row=_row, column=4).value # value of column 4 for every row
				if _progName != None:
					if str(self.sheet.cell(row=_row, column=5).value) != 'None': # if episode number not None
						_progName = self.sheet.cell(row=_row, column=4).value.partition(' (')[0] + ' Episode# - ' + str(self.sheet.cell(row=_row, column=5).value) # add episode number to program name
						programs[_progName] = {'HD':'None', 'SC':'None', 'SC_MAM':'None','DubSD':'None', 'DubHD':'None', 'DubHD_MAM':'None', 'DubSD_MAM':'None'} # build dict's from program name and episode #
						rows.append(_row) # store row num for every row used. (for later use)
					#	print('building: ', _progName)
					elif str(self.sheet.cell(row=_row, column=5).value) == 'None': # same logic as above, without episode # (it will be used for movies)
						_progName = _progName.partition(' (')[0]
						programs[_progName] = {'HD':'None', 'SC':'None', 'SC_MAM':'None','DubSD':'None', 'DubHD':'None', 'DubHD_MAM':'None', 'DubSD_MAM':'None'}
						rows.append(_row)
					#	print('building: ', _progName, 'Episode # ', str(self.sheet.cell(row=_row, column=5).value))
			print('Done buildingat shit')
			return programs, rows
		build_dict()


		def order_dict():
			print(output_file)
			print('Can i take your order please?')
			for _row in range(11, self.max_row_num): # for every row
				print('Yes, i want row #', _row, 'Please')
				if str(self.sheet.cell(row=_row, column=4).value) != 'None': # if program name not none
					print('Preparing order named', str(self.sheet.cell(row=_row, column=4).value), 'Please wait patiently')
			#		if re.search(r'(\(RUS HD)\)', str(sheet.cell(row=_row, column=4).value)):
						#print(_progName, _row)
			#			ru_row.append(_row)
			
			
					if re.search(r'(\(4K)\)', str(self.sheet.cell(row=_row, column=4).value)): # search inside parenthesis for exact phrase '4K' and skip line if found.
						print(str(self.sheet.cell(row=_row, column=5).value), "4K content, Skipped")
						pass
					elif re.search(r'(\(HD)\)', str(self.sheet.cell(row=_row, column=4).value)): # search inside parenthesis for exact phrase 'HD'
						if str(self.sheet.cell(row=_row, column=5).value) != 'None': # if episode num not none
							_progName = self.sheet.cell(row=_row, column=4).value.partition(' (')[0] + ' Episode# - ' + str(self.sheet.cell(row=_row, column=5).value) # program name with episode num
							programs[_progName]['HD'] = self.sheet.cell(row=_row, column=8).value # save value from column 8(exm; "MV859563") under the "HD" key 
							rows.append(_row)
							#for ran in range(1, 10):
								#sheet.cell(row=_row, column=ran).fill = fill_blue
							#print('HD EP', _progName)
						#	print(_row)
						elif str(self.sheet.cell(row=_row, column=4).value) != 'None': # Same logic as above without episode num in progrm name(Used for movies)
							_progName = self.sheet.cell(row=_row, column=4).value.partition(' (')[0]
							programs[_progName]['HD'] = self.sheet.cell(row=_row, column=8).value
							rows.append(_row)
							#for ran in range(1, 10):
								#sheet.cell(row=_row, column=ran).fill = fill_blue
						#	print(_row)
							#print('HD', _progName)
					
					
							
					elif re.search(r'(\(CU HD)\)', str(self.sheet.cell(row=_row, column=4).value)): # Same logic as sibling if "CU HD" inside parenthesis(used only for movies-NO EPISODE NUM in program name)
						#if str(sheet.cell(row=_row, column=8).value) != 'None':
						_progName = self.sheet.cell(row=_row, column=4).value.partition(' (')[0]
						programs[_progName]['HD'] = self.sheet.cell(row=_row, column=8).value
						rows.append(_row)
						#print(_row)
						#print('CU HD', _progName)	
						#for ran in range(1, 10):
								#sheet.cell(row=_row, column=ran).fill = fill_pink
						if self.sheet.cell(row=_row, column=9).value == 'None': # if another lackart exist for same contant
							programs[_progName]['DubHD_MAM'] = self.sheet.cell(row=_row, column=9).value
							#pass
						
						
					elif re.search(r'(\(DUB HD)\)', str(self.sheet.cell(row=_row, column=4).value)) != None: # Same logic as sibling if "DUB HD" inside parenthesis(used only for movies-NO EPISODE NUM in program name)
					#	if str(sheet.cell(row=_row, column=4).value) != 'None':
						_progName = self.sheet.cell(row=_row, column=4).value.partition(' (')[0]
						#print('DUB HD', _progName)
						programs[_progName]['DubHD'] = self.sheet.cell(row=_row, column=8).value
						rows.append(_row)
						#	if _progName == 'NORM OF THE NORTH 2' or _progName == 'YOUNG AHMED':
						#print(_progName, 'DUB HD!')
						if self.sheet.cell(row=_row, column=9).value != 'None':
							programs[_progName]['DubHD_MAM'] = self.sheet.cell(row=_row, column=9).value
						#		print('dubHD MAM')
							rows.append(_row)
							#for ran in range(1, 10):
								#sheet.cell(row=_row, column=ran).fill = fill_pink
						elif self.sheet.cell(row=_row, column=9).value == 'None':
							programs[_progName]['DubHD_MAM'] = self.sheet.cell(row=_row, column=9).value
							#pass
								
		# Rus SHIT		
		#			elif re.search(r'(\(RUS HD)\)', str(sheet.cell(row=_row, column=4).value)): # search inside parenthesis for exact phrase 'RUS HD'
		#				if str(sheet.cell(row=_row, column=5).value) != 'None': # if episode num not none
		#					_progName = sheet.cell(row=_row, column=4).value.partition(' (')[0] + ' Episode# - ' + str(sheet.cell(row=_row, column=5).value) # program name with episode num
		#					programs[_progName]['HD'] = sheet.cell(row=_row, column=8).value # save value from column 8(exm; "MV859563") under the "HD" key 
		#					rows.append(_row)
							#print('HD EP', _progName)
						#	print(_row)
		#					for ran in range(1, 10):
		#						sheet.cell(row=_row, column=ran).fill = fill_purple
		#				elif str(sheet.cell(row=_row, column=4).value) != 'None': # Same logic as above without episode num in progrm name(Used for movies)
		#					_progName = sheet.cell(row=_row, column=4).value.partition(' (')[0]
		#					programs[_progName]['HD'] = sheet.cell(row=_row, column=8).value
		#					rows.append(_row)
						#	print(_row)
		#					#print('HD', _progName)
		#					for ran in range(1, 10):
		#						sheet.cell(row=_row, column=ran).fill = fill_purple
		# END RUS SHIT			
								
					elif re.search(r'(\(DUB)\)', str(self.sheet.cell(row=_row, column=4).value)) != None: # Same logic as sibling if "DUB" inside parenthesis(used only for movies-NO EPISODE NUM in program name)
						#if str(sheet.cell(row=_row, column=5).value) != 'None':
						_progName = self.sheet.cell(row=_row, column=4).value.partition(' (')[0]
						#print('DUBSD', _progName)
						programs[_progName]['DubSD'] = self.sheet.cell(row=_row, column=8).value
						rows.append(_row)
					#	print(_row)
						if self.sheet.cell(row=_row, column=9).value != 'None':
							programs[_progName]['DubSD_MAM'] = self.sheet.cell(row=_row, column=9).value
					#		print('SDMAM`d ', _progName)
							rows.append(_row)
							#for ran in range(1, 10):
								#sheet.cell(row=_row, column=ran).fill = fill_orange
					
					elif re.search(r'(\(CU)\)', str(self.sheet.cell(row=_row, column=4).value)) != None: # Same logic as sibling if "CU" inside parenthesis(used only for movies-NO EPISODE NUM in program name)
						#if str(sheet.cell(row=_row, column=5).value) != 'None':
						_progName = self.sheet.cell(row=_row, column=4).value.partition(' (')[0]
						#print('CU', _progName)
						programs[_progName]['SC'] = self.sheet.cell(row=_row, column=8).value
						rows.append(_row)
					#	print(_row)
						if self.sheet.cell(row=_row, column=9).value != 'None':
							programs[_progName]['SC_MAM'] = self.sheet.cell(row=_row, column=9).value
					#		print('SDMAM`d ', _progName)
							rows.append(_row)
							#for ran in range(1, 10):
								#sheet.cell(row=_row, column=ran).fill = fill_blue
					
							
					elif re.search(r'(\(CU DUB)\)', str(self.sheet.cell(row=_row, column=4).value)) != None: # Same logic as sibling if "CU DUB" inside parenthesis(used only for movies-NO EPISODE NUM in program name)
						#if str(sheet.cell(row=_row, column=5).value) != 'None':
						_progName = self.sheet.cell(row=_row, column=4).value.partition(' (')[0]
						#print('DUBSD', _progName)
						programs[_progName]['DubSD'] = self.sheet.cell(row=_row, column=8).value
						rows.append(_row)
					#	print(_row)
						if self.sheet.cell(row=_row, column=9).value != 'None':
							programs[_progName]['DubSD_MAM'] = self.sheet.cell(row=_row, column=9).value
					#	print('SDMAM`d ', _progName)
							#rows.append(_row)
							#for ran in range(1, 10):
								#sheet.cell(row=_row, column=ran).fill = fill_green

					elif re.search(r'(\(CU DUB HD)\)', str(self.sheet.cell(row=_row, column=4).value)) != None: # Same logic as sibling if "CU DUB HD" inside parenthesis(used only for movies-NO EPISODE NUM in program name)
						#if str(sheet.cell(row=_row, column=9).value) == 'None':
						_progName = _progName.partition(' (')[0]
						programs[_progName]['DubHD'] = self.sheet.cell(row=_row, column=8).value
						rows.append(_row)
						#print('CU DUB HD', _progName)
						if self.sheet.cell(row=_row, column=9).value != 'None':
							programs[_progName]['DubHD_MAM'] = self.sheet.cell(row=_row, column=9).value	
						#	print('dubsd MAM - ', _progName)
							rows.append(_row)
							#for ran in range(1, 10):
								#sheet.cell(row=_row, column=ran).fill = fill_yellow
								
								
					else: # every existing dict based on program name's
						for prog in programs:
							if self.sheet.cell(row=_row, column=4).value != 'None':
								if self.sheet.cell(row=_row, column=4).value == prog:
									_progName = self.sheet.cell(row=_row, column=4).value
									programs[_progName]['SC'] = self.sheet.cell(row=_row, column=8).value
									#print('SC', _progName, _row)
									#rows.append(_row)
								#	print('sd', _row)
									if self.sheet.cell(row=_row, column=9).value != 'None':
										programs[_progName]['SC_MAM'] = self.sheet.cell(row=_row, column=9).value
										#for ran in range(1, 10):
										#	sheet.cell(row=_row, column=ran).fill = fill_red
									#else:
									#	pass
								elif self.sheet.cell(row=_row, column=4).value.partition(' (')[0] + ' Episode# - ' + str(self.sheet.cell(row=_row, column=5).value) == prog:
									_progName = self.sheet.cell(row=_row, column=4).value.partition(' (')[0] + ' Episode# - ' + str(self.sheet.cell(row=_row, column=5).value)
									programs[_progName]['SC'] = self.sheet.cell(row=_row, column=8).value
									#print('SC EP', _progName, _row)
									#rows.append(_row)
									if self.sheet.cell(row=_row, column=9).value != 'None':
										programs[_progName]['SC_MAM'] = self.sheet.cell(row=_row, column=9).value
										#for ran in range(1, 10):
										#	sheet.cell(row=_row, column=ran).fill = fill_red
									elif self.sheet.cell(row=_row, column=9).value == 'None':
										programs[_progName]['SC_MAM'] = self.sheet.cell(row=_row, column=9).value
								#else:
								#	pass

			return programs, rows
		order_dict()


		def new_file(): # Fill new xlsx files
			print('new filing tht shit')
			maxRow = self.sheet.max_row
			out_file_fixed = 'Errored File Name.xlsx'
			wb1 = Workbook()
			NewSheet = wb1.active
			NewSheet['A1'] = 'MASTER'
			NewSheet['B1'] = 'Master Type'
			NewSheet['C1'] = 'Slave'
			NewSheet['D1'] = 'Slave Type'
			NewSheet['E1'] = 'MAM XML'
			NewSheet['A1'].fill = fill_red
			NewSheet['B1'].fill = fill_red
			NewSheet['C1'].fill = fill_red
			NewSheet['D1'].fill = fill_red
			NewSheet['E1'].fill = fill_red
			NewSheet['A1'].font = red_font
			NewSheet['B1'].font = red_font
			NewSheet['C1'].font = red_font
			NewSheet['D1'].font = red_font
			NewSheet['E1'].font = red_font
			for program in programs:
				if program in programs != None:
					file_name = str(program)
					out_file = 'C:\\Users\\307311530\\Desktop\\scripts\\Zivugim\\Test\\' + file_name + '.xlsx'
					maxRow = NewSheet.max_row + 1
					hd = programs[program]['HD']
					sc = programs[program]['SC']
					MX = programs[program]['SC_MAM']
					dubsd = programs[program]['DubSD']
					dubhd = programs[program]['DubHD']
					dubsdMAM = programs[program]['DubSD_MAM']
					dubHdMAM = programs[program]['DubHD_MAM']
					if hd != 'None' and sc != 'None':
						#print(program, 'HD: ', hd, 'SC: ', sc, 'MAM SD: ', MX)
						#print('--HD/SC--', program, '--HD/SC--')
						NewSheet['A' + str(maxRow)] = hd
						#print(maxRow)
						NewSheet['B' + str(maxRow)] = 'HC'
						NewSheet['C' + str(maxRow)] = sc
						NewSheet['D' + str(maxRow)] = 'SC'
						#print(maxRow)
						#print('--HD/SC--', program, '--HD/SC--')
						if MX != 'None':
							#print(maxRow)
							NewSheet['E' + str(maxRow)] = MX
							#print('--MAM--', program, '--MAM--')
						maxRow = NewSheet.max_row + 1
						#file_name = str(program)
						#out_file = 'C:\\Users\\307311530\\Desktop\\scripts\\Zivugim\\Test\\' + file_name + '.xlsx'
						#wb1.save(out_file)
						#else:
						#	pass
					if dubsd != 'None' and hd != 'None':
						#print(program, 'HD: ', hd, 'Dub SD: ', dubsd, 'MAM DUBSD: ', dubsdMAM)
						#print('--HD/DUBSD--', program, '--HD/DUBSD--')
						NewSheet['A' + str(maxRow)] = hd
						NewSheet['B' + str(maxRow)] = 'HC'
						NewSheet['C' + str(maxRow)] = dubsd
						NewSheet['D' + str(maxRow)] = 'SDHC'
						#print(maxRow)
						#print('--HD/DUBSD--', program, '--HD/DUBSD--')
						if dubsdMAM != 'None':
							#print(maxRow)
							NewSheet['E' + str(maxRow)] = dubsdMAM
							#print('--MAM--', program, '--MAM--')
						maxRow = NewSheet.max_row + 1
						#file_name = str(program)
						#out_file = 'C:\\Users\\307311530\\Desktop\\scripts\\Zivugim\\Test\\' + file_name + '.xlsx'
						#wb1.save(out_file)
						#else:
						#	pass
					if dubhd != 'None' and hd != 'None': 
						#print(program, 'HD: ', hd, 'Dub HD: ', dubhd, 'MAM DUBHD: ', dubHdMAM)
						#print('--HD/Dub HD--', program, '--HD/Dub HD--')
						NewSheet['A' + str(maxRow)] = hd
						NewSheet['B' + str(maxRow)] = 'HC'
						NewSheet['C' + str(maxRow)] = dubhd
						NewSheet['D' + str(maxRow)] = 'HDHC'
						#print('--HD/Dub HD--', program, '--HD/Dub HD--')
						if dubHdMAM != 'None':
							#print(maxRow)
							NewSheet['E' + str(maxRow)] = dubHdMAM
							#print('--MAM--', program, '--MAM--')
						#file_name = str(program)
						#out_file = 'C:\\Users\\307311530\\Desktop\\scripts\\Zivugim\\Test\\' + file_name + '.xlsx'
						#wb1.save(out_file)
						else:
							pass
						#print('Elsed', program, 'HD: ', hd, 'HD_DUB', dubhd, 'DubHdMAM: ', dubHdMAM )
						maxRow = NewSheet.max_row + 1
						#print(maxRow)
						NewSheet['A' + str(maxRow)] = hd
						NewSheet['B' + str(maxRow)] = 'HC'
						NewSheet['C' + str(maxRow)] = dubhd
						NewSheet['D' + str(maxRow)] = 'HD_DUB'
						#print(maxRow)
						#print('--HD/Second Dub HD--', program, '--HD/Second Dub HD--')
						if dubsdMAM != 'None':
							#print(maxRow)
							NewSheet['E' + str(maxRow)] = dubHdMAM
							#print('--MAM--', program, '--MAM--')
						else:
							pass			
					if dubhd != 'None' and hd == 'None' and dubsd != 'None': 
						#print(program, 'HD: ', hd, 'Dub HD: ', dubhd, 'MAM DUBHD: ', dubHdMAM)
						#print('--HD/Dub HD--', program, '--HD/Dub HD--')
						NewSheet['A' + str(maxRow)] = dubhd
						NewSheet['B' + str(maxRow)] = 'HC'
						NewSheet['C' + str(maxRow)] = dubsd
						NewSheet['D' + str(maxRow)] = 'SC'
						#print('--HD/Dub HD--', program, '--HD/Dub HD--')
						if dubsdMAM != 'None':
							#print(maxRow)
							NewSheet['E' + str(maxRow)] = dubsdMAM
							#print('--MAM--', program, '--MAM--')
						#file_name = str(program)
						#out_file = 'C:\\Users\\307311530\\Desktop\\scripts\\Zivugim\\Test\\' + file_name + '.xlsx'
						#wb1.save(out_file)
						else:
							pass
						if dubsdMAM != 'None':
							#print(maxRow)
							NewSheet['E' + str(maxRow)] = dubsdMAM
							#print('--MAM--', program, '--MAM--')
						else:
							pass
					
			wb1.save(output_file)
		new_file()

		def log():
			print('logging that shit')
			with open('log.txt', 'w+') as f:
				#print(programs, file=f)
				for program in programs:
					_progName = program
					prog = programs[_progName]
					print(program, prog, '\n', file=f)
		log()

				
		def alert(): # Func must be after all dict`s was builded, wont work otherwise!
		# Scan dict`s and if there is HD but SD missing, alert user!
		#'''CHANGE THIS FUNC SO IT WILL FILL WITH RED ANY CONTENT LINE THAT FALL's UNDER THE IF STATMENT'''
			print('alerting for some shit')
			print('bla', self.sheet)
			for program in programs:
				if programs[program]['HD'] != 'None' and programs[program]['SC'] == 'None':
					print("Do Manualy: ", program)
			#for _row in range(11, max_row_num):
			#	if program == sheet.cell(row=_row, column=4).value:
			#		print(program)
		alert()

		outputWB = OP.load_workbook(output_file, 'W')
		outputSheet = outputWB['Sheet']
		max_row_num2 = outputSheet.max_row
		zivugim = {} # Line#NUM{'MASTER':'None', 'SLAVE':'None', 'MAM':'None"}


		def GatherLac():
			#Compare content name from input file to lackart's and their type [slave\root] in output file,
			#Build dictionery key for every line in output file and save root\slave lackart as value`s 
			#count = 0
			print('blabla', self.sheet)
			for _row in range(2, max_row_num2 +1): # for every row
				_master = outputSheet.cell(row=_row, column=1).value
				_slave = outputSheet.cell(row=_row, column=3).value
				_mam = outputSheet.cell(row=_row, column=5).value
				if outputSheet.cell(row=_row, column=2).value == 'HC' and outputSheet.cell(row=_row, column=4).value == 'SC' :
					zivugim[_row] = {'MASTER':_master, 'SLAVE':_slave, 'MAM':'None'}
					if _mam != 'None':
						zivugim[_row]['MAM'] = _mam
			return zivugim
		GatherLac()


		def ColorRows(): # loop lackart's from input file and compare to slave\root in output file,
			#Fill root's line in input file with green, fill slave line in input file with light green. root
			print('Coloring the new shit')
			for _row in range(11, self.max_row_num):
				_lakart = self.sheet.cell(row=_row, column=8).value
				print(_lakart)
				for _ziv in zivugim:
					if _lakart == 'None' or _lakart == None:
						pass
					else:
						for _cell in self.sheet.iter_cols(min_row=_row, max_row=_row, min_col=1, max_col=9):
							for __cell in _cell:
								if _lakart == zivugim[_ziv]['MASTER']:
									__cell.fill = fill_green
								elif _lakart == zivugim[_ziv]['SLAVE']:
									__cell.fill = fill_light_green
								else:
									pass
			self.wb.save(self.MagicFile)
			messagebox.showinfo("Done!", CoolExit[Coool])
		ColorRows()

#root = tk.Tk()
my_gui = ZivWindow()

if __name__ == "__main__":
	mainloop()
